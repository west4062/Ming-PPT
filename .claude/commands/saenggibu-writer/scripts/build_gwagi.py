#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
students.json  ->  교과세특 업로드 양식(.xlsx)

두 가지 모드를 자동으로 고른다.

[A] NEIS 채우기 모드 (권장 · 사용자가 NEIS 다운로드 파일을 --template 로 줄 때)
    NEIS '과목별세부능력및특기사항' 다운로드 파일은 학생개인번호·과목코드·반/번호·
    성명 등 메타데이터(A~I열)가 이미 채워져 있고 '세부능력 및 특기사항' 칸(보통 J열)만
    비어 있다. 이 모드는 그 파일을 **원본 그대로 보존**하면서, 학생을 반/번호(또는 이름)로
    매칭해 세특 칸에만 평가문을 채운다. 그대로 NEIS에 다시 올릴 수 있다.
    → --template 에 세특 열(헤더에 '세부능력 및 특기사항' 등)과 학생 데이터 행이 함께
      있으면 자동으로 이 모드가 선택된다.

[B] 새로 만들기 모드 (기존 동작 · 평평한 양식)
    기본 열: 학번 | 이름 | 교과세특 | 성적. 한 행 = 학생 한 명.
    --template 에 헤더만 있고 데이터 행이 없으면 그 헤더 순서를 따른다.
    --template 가 없으면 기본 4열.

사용법:
  # NEIS 파일에 채워 넣기 (학생코드·수업코드 보존)
  python build_gwagi.py students.json 결과_NEIS.xlsx \
      --template "2026_1학기_1학년_1_정보_과목세특_1.xlsx"

  # 새 양식으로 만들기
  python build_gwagi.py students.json 교과기초자료.xlsx
  python build_gwagi.py students.json 교과기초자료.xlsx --template "학교양식.xlsx"
"""
import sys
import re
import argparse
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# 헤더 이름 → students.json 필드 (별칭 흡수)
ALIASES = {
    "번호": "번호", "학번": "번호", "학번호": "번호",
    "이름": "이름", "성명": "이름",
    "평가문": "평가문", "교과세특": "평가문", "세특": "평가문",
    "세부능력및특기사항": "평가문", "세부능력 및 특기사항": "평가문",
    "특기사항": "평가문", "내용": "평가문",
    "성적": "성적", "점수": "성적", "등급": "성적", "총점": "성적",
    "바이트": "바이트수", "바이트수": "바이트수",
}
DEFAULT_HEADERS = ["학번", "이름", "교과세특", "성적"]

# NEIS 세특 칸으로 인식할 헤더(공백 제거 후 비교)
SETUK_HEADERS = {"세부능력및특기사항", "교과세특", "세특", "특기사항"}
# NEIS 반/번호 칸으로 인식할 헤더
BANBEON_HEADERS = {"반/번호", "반번호", "반/번", "학급/번호"}
NAME_HEADERS = {"성명", "이름"}
# NEIS 다운로드 파일임을 확정해 주는 메타데이터(학생코드·수업코드 등) 헤더
NEIS_META_HEADERS = {
    "학생개인번호", "학생코드", "개인번호",        # 학생코드
    "과목코드", "수업코드", "교과코드",            # 수업코드
    "학년도", "학적변동구분", "영재·발명교육기록사항",
}

THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def norm(s):
    return str(s).replace(" ", "").strip() if s is not None else ""


def resolve(header):
    key = norm(header)
    return ALIASES.get(str(header)) or ALIASES.get(key)


# ─────────────────────────── 매칭 키 정규화 ───────────────────────────
def parse_banbeon(value):
    """'1/12' · '1-12' · 12 · '10112'(학년반번) -> (반, 번호) 또는 (None, 번호).

    - '1/12', '1-12'           -> ('1', '12')
    - 5자리 숫자 '10112'        -> 반='1', 번호='12' (학년1·반01·번12 가정: 가운데 2 / 끝 2)
                                   * 끝 2자리=번호, 그 앞 1~2자리=반
    - 순수 번호 '12'           -> (None, '12')
    """
    if value is None:
        return (None, None)
    s = str(value).strip()
    m = re.match(r"^\s*(\d+)\s*[/\-]\s*(\d+)\s*$", s)
    if m:
        return (str(int(m.group(1))), str(int(m.group(2))))
    digits = re.sub(r"\D", "", s)
    if not digits:
        return (None, None)
    if len(digits) >= 4:
        # 끝 2자리=번호, 그 앞=반(학년 접두는 무시)
        num = str(int(digits[-2:]))
        ban = str(int(digits[-4:-2]))
        return (ban, num)
    return (None, str(int(digits)))


def student_keys(s):
    """학생 한 명에서 가능한 매칭 키들을 만든다."""
    ban, num = parse_banbeon(s.get("번호"))
    # '반' 필드가 따로 있으면 우선
    if s.get("반") not in (None, ""):
        ban = str(s.get("반")).strip()
    name = norm(s.get("이름"))
    return {"ban": ban, "num": num, "name": name}


# ─────────────────────────── NEIS 채우기 모드 ───────────────────────────
def detect_neis_template(path):
    """--template 가 NEIS형(세특 열 + 데이터 행)인지 판별. 헤더행 인덱스도 찾는다.

    return: (is_neis, header_row, col_map) 또는 (False, None, None)
      col_map = {'setuk': col, 'banbeon': col, 'name': col}
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    # 헤더행: 1~5행 중 '세특' 헤더가 있는 첫 행
    for hr in range(1, 6):
        headers = {c: norm(ws.cell(hr, c).value) for c in range(1, ws.max_column + 1)}
        setuk_col = next((c for c, h in headers.items() if h in SETUK_HEADERS), None)
        if setuk_col is None:
            continue
        banbeon_col = next((c for c, h in headers.items() if h in BANBEON_HEADERS), None)
        name_col = next((c for c, h in headers.items() if h in NAME_HEADERS), None)
        has_meta = any(h in NEIS_META_HEADERS for h in headers.values())
        # 데이터 행 수 (헤더 다음 행부터)
        data_rows = sum(
            1 for r in range(hr + 1, ws.max_row + 1)
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1))
        )
        # NEIS 확정 조건:
        #  - 학생코드/수업코드 등 메타 헤더가 있거나(가장 확실), 또는
        #  - 반/번호 칸 + 데이터 2행 이상(실제 학급 명부)
        # → 교과세특 헤더 + 예시 1행짜리 '새 양식'은 여기서 걸러져 build 모드로 감.
        is_neis = (banbeon_col or name_col) and (has_meta or (banbeon_col and data_rows >= 2))
        if is_neis:
            return (True, hr, {"setuk": setuk_col, "banbeon": banbeon_col, "name": name_col})
        # 세특 헤더는 찾았으나 NEIS 시그니처가 없으면 더 볼 필요 없음 → build 모드
        return (False, None, None)
    return (False, None, None)


def fill_neis(template, out, students):
    """NEIS 파일을 보존하며 세특 칸만 채운다."""
    is_neis, hr, cm = detect_neis_template(template)
    if not is_neis:
        return False  # 호출부에서 build 모드로 폴백

    # 스타일·메타데이터 보존을 위해 data_only=False 로 다시 연다
    wb = load_workbook(template)
    ws = wb.active

    # 학생 인덱스. (반,번호) 완전키 / 이름 / 번호-단독 을 분리해 둔다.
    # 번호-단독은 "반이 없는 양식"에서만 쓰는 마지막 보조수단이라 별도 보관한다.
    by_banbeon, by_name, by_num = {}, {}, {}
    dup_name, dup_num = set(), set()
    for s in students:
        k = student_keys(s)
        if k["ban"] and k["num"]:
            by_banbeon[(k["ban"], k["num"])] = s
        if k["name"]:
            if k["name"] in by_name:
                dup_name.add(k["name"])
            by_name[k["name"]] = s
        if k["num"]:
            if k["num"] in by_num:
                dup_num.add(k["num"])
            by_num[k["num"]] = s

    matched, unmatched_rows = [], []
    used = set()                      # 이미 배정된 학생(id) — 한 학생이 두 행에 들어가지 않게
    setuk_c = cm["setuk"]
    for r in range(hr + 1, ws.max_row + 1):
        # 빈 행 skip
        if all(ws.cell(r, c).value in (None, "") for c in range(1, ws.max_column + 1)):
            continue
        ban, num = (None, None)
        if cm["banbeon"]:
            ban, num = parse_banbeon(ws.cell(r, cm["banbeon"]).value)
        name = norm(ws.cell(r, cm["name"]).value) if cm["name"] else ""

        stu, how = None, ""
        # 1) (반,번호) 완전 매칭 — 가장 신뢰도 높음
        if ban and num and (ban, num) in by_banbeon:
            stu, how = by_banbeon[(ban, num)], "반/번호"
        # 2) 이름 매칭(중복 이름은 보류 → 사람이 확인). 느슨한 번호매칭보다 우선.
        elif name and name in by_name and name not in dup_name:
            stu, how = by_name[name], "이름"
        # 3) 번호-단독 매칭: 이 행에 '반'이 없을 때만, 중복 아닌 번호에 한해.
        elif num and not ban and num in by_num and num not in dup_num:
            stu, how = by_num[num], "번호"

        # 이미 다른 행에 배정된 학생이면 오매칭 방지를 위해 건너뛴다.
        if stu is not None and id(stu) in used:
            stu = None

        if stu is None:
            unmatched_rows.append((r, f"{ban or '?'}/{num or '?'}", name))
            continue

        text = stu.get("평가문", "") or ""
        cell = ws.cell(r, setuk_c, text)
        cell.alignment = WRAP
        cell.border = THIN
        matched.append((f"{ban or '-'}/{num or '-'}", name or stu.get("이름", ""), how))
        used.add(id(stu))

    # 세특 열 너비 넉넉히
    ws.column_dimensions[get_column_letter(setuk_c)].width = 80

    wb.save(out)

    # 리포트
    not_written = [s for s in students if id(s) not in used]
    print(f"[NEIS 채우기 모드] 저장 완료: {out}")
    print(f"  세특 열: {get_column_letter(setuk_c)}({norm(ws.cell(hr, setuk_c).value)}) "
          f"· 매칭 {len(matched)}명 / 양식 행 기준")
    if unmatched_rows:
        print(f"  ⚠ 양식에는 있으나 students.json에서 못 찾은 행 {len(unmatched_rows)}개:")
        for r, bb, nm in unmatched_rows[:20]:
            print(f"     - {r}행  반/번호={bb}  이름={nm or '(없음)'}")
    if not_written:
        print(f"  ⚠ students.json에는 있으나 양식에 안 들어간 학생 {len(not_written)}명:")
        for s in not_written[:20]:
            print(f"     - 번호={s.get('번호','?')}  이름={s.get('이름','?')}")
    if dup_name:
        print(f"  ℹ 이름 중복으로 이름매칭 보류: {', '.join(sorted(dup_name))} "
              f"(반/번호로만 매칭됨)")
    return True


# ─────────────────────────── 새로 만들기 모드 ───────────────────────────
def build_flat(template, out, students):
    if template:
        twb = load_workbook(template, data_only=True)
        tws = twb.active
        headers = [tws.cell(1, c).value for c in range(1, tws.max_column + 1)
                   if tws.cell(1, c).value is not None]
    else:
        headers = DEFAULT_HEADERS[:]

    wb = Workbook()
    ws = wb.active
    ws.title = "교과세특"

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = THIN

    for r, s in enumerate(students, 2):
        for c, h in enumerate(headers, 1):
            field = resolve(h)
            val = s.get(field, "") if field else ""
            cell = ws.cell(r, c, val)
            cell.border = THIN
            if field == "평가문":
                cell.alignment = WRAP
            elif field in ("번호", "이름", "성적", "바이트수"):
                cell.alignment = CENTER

    for c, h in enumerate(headers, 1):
        f = resolve(h)
        ws.column_dimensions[get_column_letter(c)].width = (
            80 if f == "평가문" else (10 if f in ("번호", "이름", "성적", "바이트수") else 16))
    ws.freeze_panes = "A2"

    wb.save(out)
    print(f"[새로 만들기 모드] 저장 완료: {out}  "
          f"(학생 {len(students)}명, 열: {', '.join(map(str, headers))})")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src")
    ap.add_argument("out")
    ap.add_argument("--template", help="기준 양식 .xlsx. NEIS 다운로드 파일이면 자동으로 채우기 모드.")
    ap.add_argument("--mode", choices=["auto", "neis", "build"], default="auto",
                    help="auto(기본): NEIS형이면 채우기, 아니면 새로 만들기")
    args = ap.parse_args()

    data = json.load(open(args.src, encoding="utf-8"))
    students = data.get("students", data if isinstance(data, list) else [])

    if args.mode == "build":
        build_flat(args.template, args.out, students)
        return
    if args.template and args.mode in ("auto", "neis"):
        if fill_neis(args.template, args.out, students):
            return
        if args.mode == "neis":
            sys.exit("⚠ --mode neis 인데 --template 가 NEIS형(세특 열+데이터 행)이 아닙니다.")
    build_flat(args.template, args.out, students)


if __name__ == "__main__":
    main()
