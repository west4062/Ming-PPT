#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
students.json  ->  통합 Excel(.xlsx)

한 시트에 학생별 자료를 활동별 열로 정리하고, 평가문 / 바이트수 / 목표·판정을 함께 담는다.
바이트수가 '목표바이트' 범위를 벗어나면 셀을 색으로 경고한다.

students.json 구조(예):
{
  "subject": "인공지능 기초",
  "students": [
    {
      "번호": "10101", "이름": "홍길동", "성적": 92,
      "제약사항": "결석 잦음",
      "활동": {
        "이미지분류프로젝트": "....",
        "AI서비스개발": "....",
        "자기평가": "...."
      },
      "목표바이트": "1400~1500",
      "평가문": "....",      // 4단계에서 채움
      "바이트수": 1452        // 4단계에서 채움
    }
  ]
}

사용법: python build_workbook.py students.json 결과.xlsx
"""
import sys
import os
import json

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from byte_count import count_bytes, parse_target  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl 가 필요합니다:  pip install openpyxl")

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
OVER_FILL = PatternFill("solid", fgColor="F4CCCC")   # 초과: 붉은 톤
UNDER_FILL = PatternFill("solid", fgColor="FFF2CC")  # 미달: 노란 톤
OK_FILL = PatternFill("solid", fgColor="D9EAD3")     # 적정: 녹색 톤
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def collect_activity_keys(students):
    keys = []
    for s in students:
        for k in (s.get("활동") or {}):
            if k not in keys:
                keys.append(k)
    return keys


def main():
    if len(sys.argv) < 3:
        sys.exit("사용법: python build_workbook.py students.json 결과.xlsx")
    src, out = sys.argv[1], sys.argv[2]
    with open(src, encoding="utf-8") as f:
        data = json.load(f)
    students = data.get("students", data if isinstance(data, list) else [])
    subject = data.get("subject", "")

    act_keys = collect_activity_keys(students)
    base_cols = ["번호", "이름", "성적", "제약사항"]
    tail_cols = ["목표바이트", "평가문", "바이트수", "판정"]
    headers = base_cols + act_keys + tail_cols

    wb = Workbook()
    ws = wb.active
    ws.title = (subject or "생활기록부")[:30]

    # 헤더
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN

    # 본문
    for r, s in enumerate(students, 2):
        text = s.get("평가문", "") or ""
        nbytes = s.get("바이트수") or count_bytes(text)
        lo, hi = parse_target(s.get("목표바이트", ""))
        if hi is not None and nbytes > hi:
            verdict, fill = f"초과(+{nbytes - hi})", OVER_FILL
        elif lo is not None and nbytes and nbytes < lo:
            verdict, fill = f"미달(-{lo - nbytes})", UNDER_FILL
        elif text:
            verdict, fill = "적정", OK_FILL
        else:
            verdict, fill = "", None

        row = {
            "번호": s.get("번호", ""),
            "이름": s.get("이름", ""),
            "성적": s.get("성적", ""),
            "제약사항": s.get("제약사항", ""),
            "목표바이트": s.get("목표바이트", ""),
            "평가문": text,
            "바이트수": nbytes if text else "",
            "판정": verdict,
        }
        for k in act_keys:
            row[k] = (s.get("활동") or {}).get(k, "")

        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, row.get(h, ""))
            cell.border = THIN
            if h in ("평가문",) or h in act_keys or h == "제약사항":
                cell.alignment = WRAP
            elif h in ("번호", "이름", "성적", "바이트수", "목표바이트", "판정"):
                cell.alignment = CENTER
            if h in ("바이트수", "판정") and fill is not None:
                cell.fill = fill

    # 열 너비
    widths = {"번호": 9, "이름": 9, "성적": 7, "제약사항": 18,
              "목표바이트": 12, "평가문": 70, "바이트수": 9, "판정": 11}
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 32)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    wb.save(out)
    print(f"저장 완료: {out}  (학생 {len(students)}명, 활동열 {len(act_keys)}개)")


if __name__ == "__main__":
    main()
